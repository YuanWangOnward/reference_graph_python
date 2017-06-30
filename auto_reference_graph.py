import re
import copy
import numpy as np
import csv
import functools
import pandas as pd
import subprocess
from openpyxl import load_workbook
import os
import glob

class AutoReferenceGraph:
    def __init__(self):
        """
        self.color = {'node': "#855D5D", 'label_font': "#FFFFFF", 'label': "#9B2D1F", 'content_font': "#000000",
                      'content': "#EFE7E7", 'label_emphasized': "#9B2D1F", 'content_emphasized': "#EFE7E7",
                      'label_review': "#A28E6A"}
        self.edge_style = {'Leads_to': '[ weight=4, penwidth=3, color="#855D5D"]',
                           'Cites': '[ weight=10, penwidth=2, color="#855D5D"]'}
        """
        self.color = {  # 'node': "#855D5D",
                        'node': "#595959",
                      'label_font': "#FFFFFF", 'label': "#00B0E9",
                      'content_font': "#000000", 'content': "#D0E7F8",
                      'label_emphasized': "#9B2D1F", 'content_emphasized': "#EFE7E7",
                      'label_review': "#A28E6A"}

        self.edge_style = {'Leads_to': '[ weight=4, penwidth=3, color="#BFBFBF"]',
                           'Cites': '[ weight=10, penwidth=2, color="#BFBFBF"]'}
        self.preserved_keys = ['ID', 'Title', 'Label', 'Citation', 'Note']
        self.key_replacement = {'Cited by': 'Citation'}
        self.minimal_font_size = 12
        self.maximal_font_size = 60
        self.time_line_font_size = 36

        try:
            self.template = self.load_template("./rs/nodeTemplate.txt")
        except:
            print('Node Template needs to be loaded')

    def load_template(self, template_path):
        self.template = open(template_path, 'r').read()
        return self.template

    def collect_info(self, pub, extra_info=None):
        """
        For each publication, only collect a subset info from Pulication
        :param pub: a Publication instance containing info
        :param extra_info: a string list of extra info to be collected beyond default ones
        :return: a dictionary of collected info
        """
        output = {}
        default_info = ['ID', 'author', 'year', 'title', 'citedby']
        if isinstance(extra_info, list):
            info = default_info + extra_info
        else:
            info = default_info

        for i in info:
            if i in pub.bib.keys():
                output[i] = pub.bib[i]
            elif i in dir(pub):
                output[i] = getattr(pub, i)
            else:
                output[i] = None
        return output

    def get_id(self, pub):
        """
        Get ID from pub is pub is filled, otherwise, make one which is different from the filled one.
        :param pub: a Publication instance containing info
        :return: ID string
        """
        if pub._filled is True:
            return pub.bib['ID']
        else:
            regex = re.compile('[^a-zA-Z0-9 ]')
            id_temp = regex.sub('', pub.bib['title'])
            return ''.join(id_temp.title().split(' '))

    def content_wrapper(self, text, length_perline=24):
        text = text.strip()

        # change new line mark
        text = text.replace('\n', '<BR/>')
        text = text.split('<BR/>')
        text = ' <BR/> '.join(text)

        # wrap text to certain length
        words = text.split(' ')
        words_expended = []
        character_count = 0
        for idx, word in enumerate(words):
            character_count += len(word)
            if character_count >= length_perline:
                words_expended.append(word)
                words_expended.append('<BR/>')
                character_count = 0
            else:
                words_expended.append(word)
                if '<BR/>' in word:
                    character_count = 0
        text = ' '.join(words_expended)

        # remove repeated new line mark
        text = text.replace('<BR/> <BR/>', '<BR/>')
        text = text.replace('<BR/><BR/>', '<BR/>')
        return text

    def add_a_node(self, template, items, values, display_keys=None, color=None):
        """ add a node to graph"""
        if display_keys is None:
            if hasattr(self, 'display_keys'):
                display_keys = self.display_keys
            else:
                display_keys = items

        if color is None:
            color = self.color

        tmp = copy.deepcopy(template)
        # look for label and title first, to place them on top
        # encoding label size by citation
        if "Citation" in items:
            font_size = \
                str(min((values[list(items).index("Citation")]) / 20 + self.minimal_font_size, self.maximal_font_size))
        else:
            font_size = self.minimal_font_size
        tmp = tmp.replace("$$$$",
                          '<TR><TD COLSPAN="2" BGCOLOR="colorLabel"><FONT  POINT-SIZE="' +
                          font_size + '" COLOR="colorLabelFont">'
                          + self.content_wrapper(
                              str(values[list(items).index("Label")])) + '</FONT></TD></TR>\n ' + "$$$$")
        # add title
        '''
        tmp = tmp.replace("$$$$", '<TR><TD COLSPAN="2" BGCOLOR="colorContent"><FONT COLOR="colorContentFont">'
                          + self.content_wrapper(
            str(values[list(items).index("Title")])) + '</FONT></TD></TR>\n ' + "$$$$")
        '''

        for item, value in zip(items, values):
            if item.strip() == "ID":
                tmp = tmp.replace("nodeID", value)
            elif item.strip() in ["Label", "Title", "Year", "Note", "Citation"]:
                pass
            elif item in display_keys:
                if value != None and value not in ['None']:
                    tmp = tmp.replace("$$$$",
                                      '<TR><TD COLSPAN="1" width="20">' + item + '</TD><TD COLSPAN="1" width="180">'
                                      + self.content_wrapper(str(value)) + '</TD></TR>\n ' + "$$$$")
        # note should be put at the end
        '''
        if str(values[list(items).index("Note")]) != None \
                and str(values[list(items).index("Note")]) not in ['None', 'nan']:
            tmp = tmp.replace("$$$$", '<TR><TD COLSPAN="2">'
                              + self.content_wrapper(str(values[list(items).index("Note")])) + '</TD></TR>\n ' + "$$$$")
        '''


        tmp = tmp.replace("$$$$", '')
        # tmp = tmp.replace("colorNode", self.color['node'])
        tmp = tmp.replace("colorLabelFont", self.color['label_font'])
        if "Citation" in items:
            dark = np.array([float(int(self.color['label'][1:3], 16)),
                             float(int(self.color['label'][3:5], 16)),
                             float(int(self.color['label'][5:7], 16))])
            shallow = np.array([float(int(self.color['content'][1:3], 16)),
                                float(int(self.color['content'][3:5], 16)),
                                float(int(self.color['content'][5:7], 16))])
            citation = str(int(values[list(items).index("Citation")]))
            if citation.isdigit():
                citation = float(citation)
            else:
                citation = 0.
            color_temp = dark + (64. - citation) / 64. * np.linalg.norm(shallow - dark)
            color_temp = np.minimum(color_temp, shallow)
            color_temp = np.maximum(color_temp, dark)
            # color_temp = [str(hex(int(v)))[-2:] for v in list(color_temp)]
            color_temp = [str("{:02x}".format(int(v))) for v in list(color_temp)]
            color_temp = '#' + ''.join(color_temp)
            # color_temp = self.color['label']
            tmp = tmp.replace("colorLabel", color_temp)
            # tmp = tmp.replace("colorNode", color_temp)
        else:
            tmp = tmp.replace("colorLabel", self.color['label'])
            # tmp = tmp.replace("colorNode", self.color['node'])

        tmp = tmp.replace("colorNode", self.color['node'])
        tmp = tmp.replace("colorContentFont", self.color['content_font'])
        tmp = tmp.replace("colorContent", self.color['content'])
        return tmp

    def load_relationship(self, relation_path):
        relations = list(csv.reader(open(relation_path, 'r')))
        IDs = list(set(functools.reduce(lambda x, y: x + y, [[row[0], row[1]] for row in relations])))
        df_relation = pd.DataFrame(IDs, columns=['ID'])
        year = pd.Series([str("".join(list(filter(str.isdigit, s))[:4])) for s in df_relation["ID"]])
        label = pd.Series([s[: s.find(year[i]) + 4] for i, s in enumerate(df_relation['ID'])])
        df_relation['Year'] = year
        df_relation['Label'] = label
        df_relation['if_record'] = False
        for col in df_relation.columns:
            df_relation[col] = [value.strip() if isinstance(value, str) else value for value in df_relation[col]]
        df_relation.index = df_relation['ID']
        return df_relation

    def load_reference_info(self, xlsx_path):
        """load a xlsx file for reference info"""
        wb = load_workbook(filename=xlsx_path)
        ws = wb.active
        df = pd.DataFrame(ws.values)
        df.columns = list(df.loc[0])
        df = df.drop(df.index[0])
        df.index = range(len(df.index))
        # check if year and label are in the xlsx table, if not, abstract the info from ID
        if 'year' not in [item.lower() for item in df.columns]:
            year = pd.Series([str("".join(list(filter(str.isdigit, s))[:4])) for s in df["ID"]])
            df['Year'] = year
        if 'label' not in [item.lower() for item in df.columns]:
            year = pd.Series([str("".join(list(filter(str.isdigit, s))[:4])) for s in df["ID"]])
            label = pd.Series([s[: s.find(year[i]) + 4] for i, s in enumerate(df['ID'])])
            df['Label'] = label
        for col in df.columns:
            df[col] = [value.strip() if isinstance(value, str) else value for value in df[col]]
        df['if_record'] = True
        df.index = df['ID']
        df.index = [id.strip() for id in df['ID']]

    def combine_info(self, df, df_relation):
        df = df.append(df_relation)
        df = df.drop_duplicates(subset='ID')
        return df

    def create_gv_file(self, GV_PATH, df, relations):
        with open(GV_PATH, 'w') as f:
            f.write('digraph G {\n')
            f.write('    edge [comment="Wildcard node added automatic in EG."];\n')
            f.write('    node [comment="Wildcard node added automatic in EG.",\n')
            f.write('        fontname="sans-serif"\n')
            f.write('        fontsize=' + str(self.minimal_font_size) + '];\n')
            # f.write('        size ="16, 4";\n')
            f.write('        ratio = "compress"\n')
            # f.write('        rankdir = LR;\n')
            f.write('        splines=ortho;\n')
            # f.write('        ranksep=4;\n')
            # f.write('        nodesep=0.2;\n')
            # f.write('        sep=0.3;\n')
            # add time line
            f.write('    {')
            f.write('        node[shape = plaintext fontsize = ' + str(self.time_line_font_size) + ' ];')
            for year in sorted(set(df['Year'])):
                if year != sorted(set(df['Year']))[-1]:
                    f.write('        ' + str(year) + ' ->')
                else:
                    f.write('        ' + str(year))
            f.write('    }\n')

            # add time line rank
            for year in sorted(set(df['Year'])):
                f.write('    {rank = same;')
                f.write('    ' + str(year) + ';')
                for id in df.loc[df['Year'] == year]['ID']:
                    f.write('    ' + id + ';')
                f.write('    }\n')

            # add relationship
            for row in relations:
                tmp = row
                tmp = tmp[0] + ' -> ' + tmp[1] + self.edge_style[tmp[2]]
                f.write('    ' + tmp + '\n')

            # add nodes
            for idx in df.index:
                f.write(self.add_a_node(self.template, df.columns, df.loc[idx]))
            f.write('}\n')

    def create_graph(self, gv_path, output_path, output_type):
        if output_type == 'png':
            subprocess.call(['dot', '-Tpng', gv_path, '-o', output_path])
        elif output_type == 'pdf':
            subprocess.call(['dot', '-Tpdf', gv_path, '-o', output_path])
        else:
            raise ValueError('type error')

    def make_id_and_label(self, author, year, title):
        regex = re.compile('[^a-zA-Z0-9 ]')
        temp1 = regex.sub('', author)
        temp1 = temp1.split(' ')[0]
        temp2 = str(year)
        temp3 = regex.sub('', title)
        temp3 = ''.join(temp3.title().split(' ')[0:3])
        return [''.join([temp1, temp2, temp3]), ''.join([temp1, temp2])]

    def prepare_df(self, df):
        """
        Add basic info (columns) to DataFrame, including ID, Title, Label, Note
        :param df:
        :return:
        """
        if 'ID' not in df.columns:
            for idx in df.index:
                id, label = self.make_id_and_label(df.loc[idx]['Authors'], df.loc[idx]['Year'], df.loc[idx]['Title'])
                df.set_value(idx, 'ID', id)

        if 'Label' not in df.columns:
            for idx in df.index:
                id, label = self.make_id_and_label(df.loc[idx]['Authors'], df.loc[idx]['Year'], df.loc[idx]['Title'])
                df.set_value(idx, 'Label', label)

        if 'Note' not in df.columns:
            df['Note'] = np.nan
            # for idx in df.index:
            #    df.set_value(idx, 'Note', '')
        df = df.rename(columns=self.key_replacement)
        df['Citation'] = [0 if np.isnan(i) else i for i in df['Citation']]
        # print('# reference without citation: ' + str(sum([np.isnan(i) for i in df['Citation']]))
        return df

    def remove_duplicated_relation(self, relation):
        temp = ['&'.join(i) for i in relation]
        temp = set(temp)
        return [i.split('&') for i in temp]



    def load_scupus_citation_bank(self, citation_back_path, loading_type='all'):
        """
        
        :param citation_back_path: 
        :param loading_type: 'all', 'cited_by', 'citing'
        :return: 
        """
        df_all = pd.DataFrame()
        relation_all = []
        extension = 'csv'
        os.chdir(citation_back_path)
        files = [i for i in glob.glob('*.{}'.format(extension))]
        # print(result)
        for file in files:
            print('loading ' + file)
            df = pd.read_csv(os.path.join(citation_back_path, file), sep=',')
            df = self.prepare_df(df)
            if file[-5] == '_':
                if loading_type in ['all', 'cited_by']:
                    relation_all = relation_all + [[file[:-5], id, 'Cites'] for id in df['ID']]
                    df_all = df_all.append(df)
            else:
                if loading_type in ['all', 'citing']:
                    relation_all = relation_all + [[id, file[:-4], 'Cites'] for id in df['ID']]
                    df_all = df_all.append(df)
        # remove duplicated items in relation
        relation_all = self.remove_duplicated_relation(relation_all)
        #temp = ['&'.join(i) for i in relation_all]
        #temp = set(temp)
        #relation_all = [i.split('&') for i in temp]
        df_all = df_all.drop_duplicates(subset="ID")
        df_all.index = df_all['ID']

        return [df_all, relation_all]

    def get_citation_count(self, reference_list):

        ids = set(reference_list)
        count = {}
        for id in ids:
            count[id] = reference_list.count(id)
        return count

    def find_cohesive_data_set(self, relation, n_threshold):
        """
        Give a relationshiop table, find out the set of papers that have at least n_threshold relations with other
        papers in the relationship table.
        :param relation:
        :param n_threshold:
        :return:
        """
        relation_count_temp = [[item[0], item[1]] for item in relation]
        relation_count_temp = functools.reduce(lambda x, y: x + y, relation_count_temp)

        relation_strength = pd.Series()
        for i in set(relation_count_temp):
            relation_strength[i] = relation_count_temp.count(i)
        relation_strength = relation_strength.sort_values(ascending=False)

        return relation_strength[relation_strength.values > n_threshold]

    def filter_relation_by_cohesive_set(self, relation, cohesive_set):
        """
        Find out relations between the cohesive reference set
        :param relation:
        :param cohesive_set: list like of reference ID's
        :return:
        """
        relation_cohesive = []
        for item in relation:
            if item[0] in cohesive_set and item[1] in cohesive_set:
                relation_cohesive.append(item)
        return relation_cohesive

    def filter_reference_info_by_cohesive_set(self, df, cohesive_set):
        """
        Pick the subset of pandas.DataFrame records that are in the cohesive_set
        :param df:
        :param cohesive_set:
        :return:
        """
        indexes = [idx for idx in df.index if idx in cohesive_set]

        return df.loc[indexes]

    def load_seed_collection(self, RELATION_PATH, XLSX_PATH):
        """
        Load the seed collection of references and relationship
        :param RELATION_PATH:
        :param XLSX_PATH:
        :return:
        """
        # load relationship
        relations = list(csv.reader(open(RELATION_PATH, 'r')))
        IDs = list(set(functools.reduce(lambda x, y: x + y, [[row[0], row[1]] for row in relations])))
        df_relation = pd.DataFrame(IDs, columns=['ID'])
        year = pd.Series([str("".join(list(filter(str.isdigit, s))[:4])) for s in df_relation["ID"]])
        label = pd.Series([s[: s.find(year[i]) + 4] for i, s in enumerate(df_relation['ID'])])
        df_relation['Year'] = year
        df_relation['Label'] = label
        # df_relation['if_record'] = False
        for col in df_relation.columns:
            df_relation[col] = [value.strip() if isinstance(value, str) else value for value in df_relation[col]]
        df_relation.index = df_relation['ID']

        # load in .xlsx file
        wb = load_workbook(filename=XLSX_PATH)
        ws = wb.active
        df = pd.DataFrame(ws.values)
        df.columns = list(df.loc[0])
        df = df.drop(df.index[0])
        df.index = range(len(df.index))
        # check if year and label are in the xlsx table, if not, abstract the info from ID
        if 'year' not in [item.lower() for item in df.columns]:
            year = pd.Series([int("".join(list(filter(str.isdigit, s))[:4])) for s in df["ID"]])
            df['Year'] = year
        if 'label' not in [item.lower() for item in df.columns]:
            year = pd.Series([str("".join(list(filter(str.isdigit, s))[:4])) for s in df["ID"]])
            label = pd.Series([s[: s.find(year[i]) + 4] for i, s in enumerate(df['ID'])])
            df['Label'] = label
        for col in df.columns:
            df[col] = [value.strip() if isinstance(value, str) else value for value in df[col]]
        # df['if_record'] = True
        df.index = df['ID']
        df.index = [id.strip() for id in df['ID']]

        # combine information
        df = df.append(df_relation)
        df = df.drop_duplicates(subset='ID')

        return [df, relations]

    def relation_to_ids(self, relations):
        IDs = list(set(functools.reduce(lambda x, y: x + y, [[row[0], row[1]] for row in relations])))
        return IDs

    def remove_isolated_reference(self, df, relations):
        """
        remove isolated item in reference list, namly without any relationship
        :param df:
        :param relations:
        :return:
        """
        id_relation = set(self.relation_to_ids(relations))
        id_df = set(df.index)
        ids = id_relation.intersection(id_df)
        return df.loc[ids]





